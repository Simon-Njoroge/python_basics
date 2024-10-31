import openpyxl as xl 
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook("mysheet.xlsx")
sheet=wb['Sheet1']

for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    correct_age=cell.value+5
    corrected_cell=sheet.cell(row,6)
    corrected_cell.value=correct_age

values=Reference(sheet,min_row=2,
          max_row=sheet.max_row,
          min_col=6,
          max_col=6)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,"g2")
wb.save("mysheet3.xlsx")

